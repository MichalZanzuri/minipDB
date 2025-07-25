import tkinter as tk
from tkinter import messagebox, ttk, filedialog
from datetime import datetime
from db import connect
import datetime as dt
import os

# ייבואים עבור Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# צבעים
BG_COLOR = "#f8fafc"
TITLE_COLOR = "#1e293b"
SUBTITLE_COLOR = "#64748b"

# צבעי כפתורים
BUTTON_COLORS = {
    "products": "#3b82f6",
    "sales": "#10b981",
    "stats": "#8b5cf6",
    "discounts": "#f59e0b",
    "queries": "#6366f1"
}

class ExcelExporter:
    """מחלקה לטיפול ביצוא נתונים לאקסל"""

    def __init__(self):
        self.workbook = None
        self.worksheet = None

    def create_styled_workbook(self, title="דוח מערכת ניהול החנות"):
        """יצירת קובץ אקסל עם עיצוב מקצועי"""
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "דוח נתונים"

        # הגדרת סגנונות
        self.header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        self.title_font = Font(name='Calibri', size=16, bold=True, color='1F4E79')
        self.data_font = Font(name='Calibri', size=11)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        return self.workbook

    def add_title_and_metadata(self, title, sheet_name="דוח נתונים"):
        """הוספת כותרת ומטא-דטה"""
        if not self.worksheet:
            return

        # כותרת ראשית
        self.worksheet['A1'] = title
        self.worksheet['A1'].font = self.title_font
        self.worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')

        # תאריך ושעה
        current_time = datetime.now().strftime("%d/%m/%Y %H:%M")
        self.worksheet['A2'] = f"נוצר בתאריך: {current_time}"
        self.worksheet['A2'].font = Font(name='Calibri', size=10, italic=True)

        # מיזוג תאים לכותרת
        self.worksheet.merge_cells('A1:F1')

    def export_data_to_excel(self, data, headers, title, filename=None):
        """יצוא נתונים לאקסל עם עיצוב מקצועי"""
        try:
            self.create_styled_workbook(title)
            self.add_title_and_metadata(title)

            # התחלת נתונים משורה 4
            start_row = 4

            # הוספת כותרות עמודות
            for col_num, header in enumerate(headers, 1):
                cell = self.worksheet.cell(row=start_row, column=col_num)
                cell.value = header
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.border

            # הוספת נתונים
            for row_num, row_data in enumerate(data, start_row + 1):
                for col_num, value in enumerate(row_data, 1):
                    cell = self.worksheet.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.font = self.data_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = self.border

            # התאמת רוחב עמודות
            for col_num in range(1, len(headers) + 1):
                column_letter = get_column_letter(col_num)
                self.worksheet.column_dimensions[column_letter].width = 15

            # שמירת הקובץ
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"דוח_{title.replace(' ', '_')}_{timestamp}.xlsx"

            # בחירת מיקום שמירה
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=filename,
                title="שמור קובץ אקסל"
            )

            if file_path:
                self.workbook.save(file_path)
                return file_path

            return None

        except Exception as e:
            raise Exception(f"שגיאה ביצוא לאקסל: {str(e)}")

class MainApplication:
    def __init__(self, root):
        self.root = root
        self.root.title("מערכת ניהול חנות")
        self.root.geometry("1920x1080")
        self.root.configure(bg=BG_COLOR)

        # הוספת ExcelExporter
        self.excel_exporter = ExcelExporter()

        self.create_main_layout()

    def create_main_layout(self):
        # מסגרת ראשית
        self.main_frame = tk.Frame(self.root, bg=BG_COLOR)
        self.main_frame.pack(fill=tk.BOTH, expand=True, padx=40, pady=40)

        # כותרת ראשית
        self.title_frame = tk.Frame(self.main_frame, bg=BG_COLOR)
        self.title_frame.pack(anchor="ne", pady=(0, 40))

        # תאריך ושעה מעל הכותרת
        self.datetime_header = tk.Frame(self.title_frame, bg=BG_COLOR)
        self.datetime_header.pack(anchor="e", pady=(0, 15))

        self.datetime_label = tk.Label(
            self.datetime_header,
            text="",
            font=("Segoe UI", 12, "bold"),
            fg="#4f46e5",
            bg=BG_COLOR,
            anchor="e"
        )
        self.datetime_label.pack(anchor="e")

        # הפעלת עדכון התאריך והשעה
        self.setup_datetime()

        tk.Label(
            self.title_frame,
            text="מערכת ניהול חנות מקצועית",
            font=("Segoe UI", 32, "bold"),
            fg=TITLE_COLOR,
            bg=BG_COLOR,
            anchor="e"
        ).pack(anchor="e")

        # מסגרת תוכן ראשית
        self.content_frame = tk.Frame(self.main_frame, bg=BG_COLOR)
        self.content_frame.pack(fill=tk.BOTH, expand=True)

        # מסגרת לכפתורים בצד ימין
        self.buttons_frame = tk.Frame(self.content_frame, bg=BG_COLOR)
        self.buttons_frame.pack(side=tk.RIGHT, anchor="ne", padx=(0, 80))

        # מסגרת תוכן דינמי בצד שמאל
        self.dynamic_content = tk.Frame(self.content_frame, bg="white", relief="solid", bd=2)
        self.dynamic_content.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 50))

        # יצירת הכפתורים
        self.create_buttons()

        # הצגת מסך ברירת מחדל
        self.show_welcome_screen()

    def setup_datetime(self):
        """הגדרת עדכון תאריך ושעה"""
        def update_datetime():
            now = datetime.now()
            date_str = now.strftime("%d-%m-%Y")
            time_str = now.strftime("%H:%M:%S")
            self.datetime_label.config(text=f"📅 {date_str} | 🕐 {time_str}")
            self.root.after(1000, update_datetime)

        update_datetime()

    def create_buttons(self):
        """יצירת כפתורי הניווט"""
        def create_styled_button(parent, text, icon, command, color):
            btn_frame = tk.Frame(parent, bg=BG_COLOR)
            btn_frame.pack(pady=15, anchor="e")

            button = tk.Button(
                btn_frame,
                text=f"{icon}  {text}",
                command=command,
                font=("Segoe UI", 14, "bold"),
                bg=color,
                fg="white",
                width=25,
                height=2,
                relief="flat",
                cursor="hand2",
                bd=0,
                highlightthickness=0
            )
            button.pack(side=tk.RIGHT)

            def on_enter(e):
                button.config(bg=self.adjust_color(color, -15))

            def on_leave(e):
                button.config(bg=color)

            button.bind("<Enter>", on_enter)
            button.bind("<Leave>", on_leave)

            return button

        # יצירת הכפתורים
        create_styled_button(self.buttons_frame, "ניהול מוצרים", "📦", self.show_products_screen, BUTTON_COLORS["products"])
        create_styled_button(self.buttons_frame, "ניהול מכירות", "🛒", self.show_sales_screen, BUTTON_COLORS["sales"])
        create_styled_button(self.buttons_frame, "ניהול לקוחות", "👥", self.show_customers_screen, "#8b5cf6")
        create_styled_button(self.buttons_frame, "סטטיסטיקות", "📊", self.show_stats_screen, BUTTON_COLORS["stats"])
        create_styled_button(self.buttons_frame, "ניהול הנחות", "💸", self.show_discounts_screen, BUTTON_COLORS["discounts"])
        create_styled_button(self.buttons_frame, "שאילתות", "⚙️", self.show_queries_screen, BUTTON_COLORS["queries"])

    def adjust_color(self, color, adjustment):
        """התאמת צבע (להכהה/להבהיר)"""
        color = color.lstrip('#')
        rgb = tuple(int(color[i:i+2], 16) for i in (0, 2, 4))
        adjusted_rgb = tuple(max(0, min(255, c + adjustment)) for c in rgb)
        return f"#{adjusted_rgb[0]:02x}{adjusted_rgb[1]:02x}{adjusted_rgb[2]:02x}"

    def clear_content(self):
        """ניקוי התוכן הדינמי"""
        for widget in self.dynamic_content.winfo_children():
            widget.destroy()

    def create_export_buttons_frame(self, parent):
        """יצירת מסגרת כפתורי יצוא"""
        export_frame = tk.Frame(parent, bg="#f8fafc", relief="solid", bd=1)
        export_frame.pack(fill=tk.X, padx=20, pady=10)

        tk.Label(
            export_frame,
            text="📊 יצוא נתונים לאקסל",
            font=("Segoe UI", 14, "bold"),
            bg="#f8fafc",
            fg="#1e293b"
        ).pack(pady=10)

        buttons_container = tk.Frame(export_frame, bg="#f8fafc")
        buttons_container.pack(pady=10)

        return export_frame, buttons_container

    def create_export_button(self, parent, text, command, icon="📤", bg_color="#059669"):
        """יצירת כפתור יצוא מעוצב"""
        btn = tk.Button(
            parent,
            text=f"{icon} {text}",
            command=command,
            font=("Segoe UI", 10, "bold"),
            bg=bg_color,
            fg="white",
            width=20,
            height=2,
            relief="flat",
            cursor="hand2"
        )
        btn.pack(side=tk.RIGHT, padx=5)

        def on_enter(e):
            btn.config(bg=self.adjust_color(bg_color, -15))
        def on_leave(e):
            btn.config(bg=bg_color)

        btn.bind("<Enter>", on_enter)
        btn.bind("<Leave>", on_leave)

        return btn

    # פונקציות יצוא לאקסל
    def export_products_to_excel(self):
        """יצוא מוצרים לאקסל"""
        try:
            conn = connect()
            cur = conn.cursor()
            cur.execute("""
                SELECT product_id, product_name, price, amount, category, 
                       minamount, added_date, last_updated 
                FROM product 
                ORDER BY product_name
            """)

            data = cur.fetchall()
            headers = ["מספר מוצר", "שם מוצר", "מחיר", "כמות", "קטגוריה",
                       "כמות מינימלית", "תאריך הוספה", "עדכון אחרון"]

            file_path = self.excel_exporter.export_data_to_excel(
                data, headers, "רשימת מוצרים"
            )

            if file_path:
                messagebox.showinfo("הצלחה", f"הקובץ נשמר בהצלחה:\n{file_path}")
                os.startfile(os.path.dirname(file_path))

            cur.close()
            conn.close()

        except Exception as e:
            messagebox.showerror("שגיאה", f"שגיאה ביצוא מוצרים: {str(e)}")

    def export_sales_to_excel(self):
        """יצוא מכירות לאקסל"""
        try:
            conn = connect()
            cur = conn.cursor()
            cur.execute("""
                SELECT saleid, saledate, totalprice, customerid
                FROM sale 
                ORDER BY saledate DESC
            """)

            data = cur.fetchall()
            headers = ["מספר מכירה", "תאריך", "סכום כולל", "קוד לקוח"]

            file_path = self.excel_exporter.export_data_to_excel(
                data, headers, "רשימת מכירות"
            )

            if file_path:
                messagebox.showinfo("הצלחה", f"הקובץ נשמר בהצלחה:\n{file_path}")
                os.startfile(os.path.dirname(file_path))

            cur.close()
            conn.close()

        except Exception as e:
            messagebox.showerror("שגיאה", f"שגיאה ביצוא מכירות: {str(e)}")

    def export_low_stock_to_excel(self):
        """יצוא מוצרים עם מלאי נמוך לאקסל"""
        try:
            conn = connect()
            cur = conn.cursor()
            cur.execute("""
                SELECT product_id, product_name, amount, minamount, 
                       (minamount - amount) as shortage, category
                FROM product 
                WHERE amount < minamount
                ORDER BY (minamount - amount) DESC
            """)

            data = cur.fetchall()
            headers = ["מספר מוצר", "שם מוצר", "כמות נוכחית", "כמות מינימלית",
                       "חסר", "קטגוריה"]

            file_path = self.excel_exporter.export_data_to_excel(
                data, headers, "מוצרים עם מלאי נמוך"
            )

            if file_path:
                messagebox.showinfo("הצלחה", f"הקובץ נשמר בהצלחה:\n{file_path}")
                os.startfile(os.path.dirname(file_path))

            cur.close()
            conn.close()

        except Exception as e:
            messagebox.showerror("שגיאה", f"שגיאה ביצוא מלאי נמוך: {str(e)}")

    def export_discounts_to_excel(self):
        """יצוא הנחות לאקסל"""
        try:
            conn = connect()
            cur = conn.cursor()
            cur.execute("""
                SELECT d.discountid, p.product_name, s.storelocation, 
                       d.discountrate, d.startdate, d.enddate,
                       CASE 
                           WHEN d.startdate <= CURRENT_DATE AND d.enddate >= CURRENT_DATE 
                           THEN 'פעילה' 
                           ELSE 'לא פעילה' 
                       END as status
                FROM discount d
                JOIN product p ON d.productid = p.product_id
                JOIN store s ON d.storeid = s.storeid
                ORDER BY d.discountid DESC
            """)

            data = cur.fetchall()
            headers = ["מספר הנחה", "מוצר", "סניף", "שיעור הנחה (%)",
                       "תאריך התחלה", "תאריך סיום", "סטטוס"]

            file_path = self.excel_exporter.export_data_to_excel(
                data, headers, "רשימת הנחות"
            )

            if file_path:
                messagebox.showinfo("הצלחה", f"הקובץ נשמר בהצלחה:\n{file_path}")
                os.startfile(os.path.dirname(file_path))

            cur.close()
            conn.close()

        except Exception as e:
            messagebox.showerror("שגיאה", f"שגיאה ביצוא הנחות: {str(e)}")

    def update_sale(self):
        """עדכון מכירה קיימת"""
        try:
            sale_id = self.sale_id_entry.get().strip()
            if not sale_id:
                messagebox.showwarning("אזהרה", "אנא הכנס מספר מכירה לעדכון")
                return

            # בדיקה שהמכירה קיימת
            conn = connect()
            cur = conn.cursor()
            cur.execute("SELECT * FROM sale WHERE saleid = %s", (int(sale_id),))
            if not cur.fetchone():
                messagebox.showerror("שגיאה", "מכירה זו לא קיימת במערכת")
                cur.close()
                conn.close()
                return

            # עדכון המכירה
            cur.execute("""
                UPDATE sale 
                SET totalprice = %s, customerid = %s
                WHERE saleid = %s
            """, (
                float(self.sale_total_entry.get()),
                int(self.sale_customer_entry.get()),
                int(sale_id)
            ))

            conn.commit()
            cur.close()
            conn.close()

            messagebox.showinfo("הצלחה", "המכירה עודכנה בהצלחה!")

            # ניקוי השדות
            self.sale_id_entry.delete(0, tk.END)
            self.sale_total_entry.delete(0, tk.END)
            self.sale_customer_entry.delete(0, tk.END)

            self.show_sales_screen()

        except ValueError:
            messagebox.showerror("שגיאה", "אנא הכנס ערכים תקינים")
        except Exception as e:
            messagebox.showerror("שגיאה", f"שגיאה בעדכון המכירה: {str(e)}")

    def delete_sale(self):
        """מחיקת מכירה"""
        try:
            sale_id = self.sale_id_entry.get().strip()
            if not sale_id:
                messagebox.showwarning("אזהרה", "אנא הכנס מספר מכירה למחיקה")
                return

            # אישור מחיקה
            result = messagebox.askyesno(
                "אישור מחיקה",
                f"האם אתה בטוח שברצונך למחוק את המכירה מספר {sale_id}?\nפעולה זו לא ניתנת לביטול!"
            )

            if not result:
                return

            conn = connect()
            cur = conn.cursor()

            # בדיקה שהמכירה קיימת
            cur.execute("SELECT * FROM sale WHERE saleid = %s", (int(sale_id),))
            if not cur.fetchone():
                messagebox.showerror("שגיאה", "מכירה זו לא קיימת במערכת")
                cur.close()
                conn.close()
                return

            # מחיקת המכירה
            cur.execute("DELETE FROM sale WHERE saleid = %s", (int(sale_id),))
            conn.commit()
            cur.close()
            conn.close()

            messagebox.showinfo("הצלחה", "המכירה נמחקה בהצלחה!")

            # ניקוי השדות
            self.sale_id_entry.delete(0, tk.END)
            self.sale_total_entry.delete(0, tk.END)
            self.sale_customer_entry.delete(0, tk.END)

            self.show_sales_screen()

        except ValueError:
            messagebox.showerror("שגיאה", "אנא הכנס מספר מכירה תקין")
        except Exception as e:
            messagebox.showerror("שגיאה", f"שגיאה במחיקת המכירה: {str(e)}")

    def show_customers_screen(self):
        """הצגת מסך ניהול לקוחות"""
        self.clear_content()

        # כותרת
        header = tk.Frame(self.dynamic_content, bg="#8b5cf6", height=50)
        header.pack(fill=tk.X)
        header.pack_propagate(False)

        tk.Label(
            header,
            text="👥 ניהול לקוחות",
            font=("Segoe UI", 16, "bold"),
            bg="#8b5cf6",
            fg="white"
        ).pack(expand=True)

        # תוכן עם גלילה
        canvas = tk.Canvas(self.dynamic_content, bg="white")
        scrollbar = ttk.Scrollbar(self.dynamic_content, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg="white")

        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # טופס ניהול לקוחות
        form_frame = tk.Frame(scrollable_frame, bg="#f8fafc", relief="solid", bd=1)
        form_frame.pack(fill=tk.X, padx=20, pady=10)

        tk.Label(form_frame, text="👤 ניהול לקוחות", font=("Segoe UI", 16, "bold"), bg="#f8fafc", fg="#1e293b").pack(pady=15)

        # שדות הטופס
        fields_frame = tk.Frame(form_frame, bg="#f8fafc")
        fields_frame.pack(padx=20, pady=10)

        def create_customer_field(parent, label_text, icon, width=15):
            row = tk.Frame(parent, bg="#f8fafc")
            row.pack(fill=tk.X, pady=8)
            tk.Label(row, text=f"{icon} {label_text}", bg="#f8fafc", font=("Segoe UI", 10, "bold"), width=18, anchor="e").pack(side=tk.RIGHT, padx=5)
            entry = tk.Entry(row, width=width, font=("Segoe UI", 11), justify="center")
            entry.pack(side=tk.RIGHT, padx=5)
            return entry

        self.customer_id_entry = create_customer_field(fields_frame, "מספר לקוח (לעדכון)", "🔢")
        self.customer_name_entry = create_customer_field(fields_frame, "שם לקוח", "👤")
        self.customer_phone_entry = create_customer_field(fields_frame, "טלפון", "📞")
        self.customer_email_entry = create_customer_field(fields_frame, "אימייל", "📧")
        self.customer_address_entry = create_customer_field(fields_frame, "כתובת", "🏠")

        # כפתורי פעולות לקוחות
        buttons_frame = tk.Frame(form_frame, bg="#f8fafc")
        buttons_frame.pack(pady=15)

        def add_customer():
            try:
                conn = connect()
                cur = conn.cursor()
                cur.execute("""
                    INSERT INTO customer (customer_name, phone, email, address)
                    VALUES (%s, %s, %s, %s)
                """, (
                    self.customer_name_entry.get(),
                    self.customer_phone_entry.get(),
                    self.customer_email_entry.get(),
                    self.customer_address_entry.get()
                ))
                conn.commit()
                cur.close()
                conn.close()
                messagebox.showinfo("הצלחה", "הלקוח נוסף בהצלחה!")
                self.clear_customer_fields()
                self.show_customers_screen()
            except Exception as e:
                messagebox.showerror("שגיאה", f"שגיאה בהוספת לקוח: {str(e)}")

        def update_customer():
            try:
                customer_id = self.customer_id_entry.get().strip()
                if not customer_id:
                    messagebox.showwarning("אזהרה", "אנא הכנס מספר לקוח לעדכון")
                    return

                conn = connect()
                cur = conn.cursor()
                cur.execute("SELECT * FROM customer WHERE customerid = %s", (int(customer_id),))
                if not cur.fetchone():
                    messagebox.showerror("שגיאה", "לקוח זה לא קיים במערכת")
                    cur.close()
                    conn.close()
                    return

                cur.execute("""
                    UPDATE customer 
                    SET customer_name = %s, phone = %s, email = %s, address = %s
                    WHERE customerid = %s
                """, (
                    self.customer_name_entry.get(),
                    self.customer_phone_entry.get(),
                    self.customer_email_entry.get(),
                    self.customer_address_entry.get(),
                    int(customer_id)
                ))

                conn.commit()
                cur.close()
                conn.close()

                messagebox.showinfo("הצלחה", "הלקוח עודכן בהצלחה!")
                self.clear_customer_fields()
                self.show_customers_screen()

            except ValueError:
                messagebox.showerror("שגיאה", "אנא הכנס מספר לקוח תקין")
            except Exception as e:
                messagebox.showerror("שגיאה", f"שגיאה בעדכון הלקוח: {str(e)}")

        def delete_customer():
            try:
                customer_id = self.customer_id_entry.get().strip()
                if not customer_id:
                    messagebox.showwarning("אזהרה", "אנא הכנס מספר לקוח למחיקה")
                    return

                result = messagebox.askyesno(
                    "אישור מחיקה",
                    f"האם אתה בטוח שברצונך למחוק את הלקוח מספר {customer_id}?\nפעולה זו לא ניתנת לביטול!"
                )

                if not result:
                    return

                conn = connect()
                cur = conn.cursor()

                cur.execute("SELECT * FROM customer WHERE customerid = %s", (int(customer_id),))
                if not cur.fetchone():
                    messagebox.showerror("שגיאה", "לקוח זה לא קיים במערכת")
                    cur.close()
                    conn.close()
                    return

                cur.execute("DELETE FROM customer WHERE customerid = %s", (int(customer_id),))
                conn.commit()
                cur.close()
                conn.close()

                messagebox.showinfo("הצלחה", "הלקוח נמחק בהצלחה!")
                self.clear_customer_fields()
                self.show_customers_screen()

            except ValueError:
                messagebox.showerror("שגיאה", "אנא הכנס מספר לקוח תקין")
            except Exception as e:
                messagebox.showerror("שגיאה", f"שגיאה במחיקת הלקוח: {str(e)}")

        tk.Button(buttons_frame, text="➕ הוסף לקוח", command=add_customer, bg="#8b5cf6", fg="white", font=("Segoe UI", 10, "bold"), width=15, height=2).pack(side=tk.LEFT, padx=5)
        tk.Button(buttons_frame, text="✏️ עדכן לקוח", command=update_customer, bg="#3b82f6", fg="white", font=("Segoe UI", 10, "bold"), width=15, height=2).pack(side=tk.LEFT, padx=5)
        tk.Button(buttons_frame, text="🗑️ מחק לקוח", command=delete_customer, bg="#ef4444", fg="white", font=("Segoe UI", 10, "bold"), width=15, height=2).pack(side=tk.LEFT, padx=5)

        # כפתורי יצוא לאקסל
        export_frame, buttons_container = self.create_export_buttons_frame(scrollable_frame)

        self.create_export_button(
            buttons_container,
            "יצא רשימת לקוחות",
            self.export_customers_to_excel,
            "👥",
            "#8b5cf6"
        )

        # טבלת לקוחות
        table_frame = tk.Frame(scrollable_frame, bg="white")
        table_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        tk.Label(table_frame, text="📋 רשימת לקוחות:", font=("Segoe UI", 14, "bold"), bg="white", fg="#1e293b").pack(anchor="e", pady=(0, 10))

        columns = ("מספר", "שם", "טלפון", "אימייל", "כתובת")
        tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=10)

        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=120, anchor="center")

        tree_scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=tree_scrollbar.set)

        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        tree_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # טעינת לקוחות
        try:
            conn = connect()
            cur = conn.cursor()
            cur.execute("SELECT customerid, customer_name, phone, email, address FROM customer ORDER BY customerid DESC LIMIT 20")
            for row in cur.fetchall():
                tree.insert("", tk.END, values=row)
            cur.close()
            conn.close()
        except Exception as e:
            pass

    def clear_customer_fields(self):
        """ניקוי שדות הלקוח"""
        self.customer_id_entry.delete(0, tk.END)
        self.customer_name_entry.delete(0, tk.END)
        self.customer_phone_entry.delete(0, tk.END)
        self.customer_email_entry.delete(0, tk.END)
        self.customer_address_entry.delete(0, tk.END)

    def export_customers_to_excel(self):
        """יצוא לקוחות לאקסל"""
        try:
            conn = connect()
            cur = conn.cursor()
            cur.execute("""
                SELECT customerid, customer_name, phone, email, address
                FROM customer 
                ORDER BY customer_name
            """)

            data = cur.fetchall()
            headers = ["מספר לקוח", "שם לקוח", "טלפון", "אימייל", "כתובת"]

            file_path = self.excel_exporter.export_data_to_excel(
                data, headers, "רשימת לקוחות"
            )

            if file_path:
                messagebox.showinfo("הצלחה", f"הקובץ נשמר בהצלחה:\n{file_path}")
                os.startfile(os.path.dirname(file_path))

            cur.close()
            conn.close()

        except Exception as e:
            messagebox.showerror("שגיאה", f"שגיאה ביצוא לקוחות: {str(e)}")

    def update_product(self):
        """עדכון מוצר קיים"""
        try:
            product_id = self.product_id_entry.get().strip()
            if not product_id:
                messagebox.showwarning("אזהרה", "אנא הכנס מספר מוצר לעדכון")
                return

            # בדיקה שהמוצר קיים
            conn = connect()
            cur = conn.cursor()
            cur.execute("SELECT * FROM product WHERE product_id = %s", (int(product_id),))
            if not cur.fetchone():
                messagebox.showerror("שגיאה", "מוצר זה לא קיים במערכת")
                cur.close()
                conn.close()
                return

            # עדכון המוצר
            cur.execute("""
                UPDATE product 
                SET product_name = %s, price = %s, amount = %s, category = %s, 
                    minamount = %s, last_updated = %s
                WHERE product_id = %s
            """, (
                self.product_name_entry.get(),
                float(self.product_price_entry.get()),
                int(self.product_amount_entry.get()),
                self.product_category_entry.get(),
                int(self.product_min_entry.get()),
                dt.date.today(),
                int(product_id)
            ))

            conn.commit()
            cur.close()
            conn.close()

            messagebox.showinfo("הצלחה", "המוצר עודכן בהצלחה!")
            self.clear_product_fields()
            self.show_products_screen()

        except ValueError:
            messagebox.showerror("שגיאה", "אנא הכנס ערכים תקינים")
        except Exception as e:
            messagebox.showerror("שגיאה", f"שגיאה בעדכון המוצר: {str(e)}")

    def delete_product(self):
        """מחיקת מוצר"""
        try:
            product_id = self.product_id_entry.get().strip()
            if not product_id:
                messagebox.showwarning("אזהרה", "אנא הכנס מספר מוצר למחיקה")
                return

            # אישור מחיקה
            result = messagebox.askyesno(
                "אישור מחיקה",
                f"האם אתה בטוח שברצונך למחוק את המוצר מספר {product_id}?\nפעולה זו לא ניתנת לביטול!"
            )

            if not result:
                return

            conn = connect()
            cur = conn.cursor()

            # בדיקה שהמוצר קיים
            cur.execute("SELECT * FROM product WHERE product_id = %s", (int(product_id),))
            if not cur.fetchone():
                messagebox.showerror("שגיאה", "מוצר זה לא קיים במערכת")
                cur.close()
                conn.close()
                return

            # מחיקת המוצר
            cur.execute("DELETE FROM product WHERE product_id = %s", (int(product_id),))
            conn.commit()
            cur.close()
            conn.close()

            messagebox.showinfo("הצלחה", "המוצר נמחק בהצלחה!")
            self.clear_product_fields()
            self.show_products_screen()

        except ValueError:
            messagebox.showerror("שגיאה", "אנא הכנס מספר מוצר תקין")
        except Exception as e:
            messagebox.showerror("שגיאה", f"שגיאה במחיקת המוצר: {str(e)}")

    def clear_product_fields(self):
        """ניקוי שדות המוצר"""
        self.product_id_entry.delete(0, tk.END)
        self.product_name_entry.delete(0, tk.END)
        self.product_price_entry.delete(0, tk.END)
        self.product_amount_entry.delete(0, tk.END)
        self.product_category_entry.delete(0, tk.END)
        self.product_min_entry.delete(0, tk.END)

    def show_low_stock_window(self):
        """הצגת חלון מוצרים עם מלאי נמוך"""
        low_stock_window = tk.Toplevel(self.root)
        low_stock_window.title("מוצרים עם מלאי נמוך")
        low_stock_window.geometry("1000x600")
        low_stock_window.configure(bg="white")

        # כותרת
        header_frame = tk.Frame(low_stock_window, bg="#ef4444", height=60)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)

        tk.Label(
            header_frame,
            text="⚠️ מוצרים עם מלאי נמוך - דורש טיפול מיידי!",
            font=("Segoe UI", 16, "bold"),
            bg="#ef4444",
            fg="white"
        ).pack(expand=True)

        # כפתור יצוא
        export_frame = tk.Frame(low_stock_window, bg="white")
        export_frame.pack(fill=tk.X, padx=20, pady=10)

        tk.Button(
            export_frame,
            text="📤 יצא לאקסל",
            command=self.export_low_stock_to_excel,
            bg="#059669",
            fg="white",
            font=("Segoe UI", 12, "bold"),
            width=15,
            height=2
        ).pack(side=tk.RIGHT)

        # טבלת נתונים
        tree_frame = tk.Frame(low_stock_window, bg="white")
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        columns = ("מספר מוצר", "שם מוצר", "כמות נוכחית", "כמות מינימלית", "חסר", "קטגוריה")
        tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=20)

        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=150, anchor="center")

        # צבעים לחומרת הבעיה
        tree.tag_configure("critical", background="#fef2f2", foreground="#dc2626")
        tree.tag_configure("warning", background="#fff8e1", foreground="#f59e0b")

        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)

        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # טעינת נתונים
        try:
            conn = connect()
            cur = conn.cursor()
            cur.execute("""
                SELECT product_id, product_name, amount, minamount, 
                       (minamount - amount) as shortage, category
                FROM product 
                WHERE amount < minamount
                ORDER BY (minamount - amount) DESC
            """)

            rows = cur.fetchall()
            for row in rows:
                shortage = row[4]
                # קביעת צבע לפי חומרת המחסור
                tag = "critical" if shortage >= 10 else "warning"
                tree.insert("", tk.END, values=row, tags=(tag,))

            cur.close()
            conn.close()

            # הצגת סיכום
            summary_frame = tk.Frame(low_stock_window, bg="#f8fafc", relief="solid", bd=1)
            summary_frame.pack(fill=tk.X, padx=20, pady=10)

            tk.Label(
                summary_frame,
                text=f"📊 סה\"כ נמצאו {len(rows)} מוצרים עם מלאי נמוך",
                font=("Segoe UI", 12, "bold"),
                bg="#f8fafc",
                fg="#dc2626"
            ).pack(pady=10)

        except Exception as e:
            tk.Label(tree_frame, text=f"שגיאה בטעינת נתונים: {e}", bg="white", fg="red").pack()

    def show_unique_customers_window(self):
        """הצגת חלון לקוחות ייחודיים"""
        customers_window = tk.Toplevel(self.root)
        customers_window.title("לקוחות ייחודיים")
        customers_window.geometry("1000x600")
        customers_window.configure(bg="white")

        # כותרת
        header_frame = tk.Frame(customers_window, bg="#8b5cf6", height=60)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)

        tk.Label(
            header_frame,
            text="👥 לקוחות ייחודיים - מי קנה במערכת",
            font=("Segoe UI", 16, "bold"),
            bg="#8b5cf6",
            fg="white"
        ).pack(expand=True)

        # כפתור יצוא
        export_frame = tk.Frame(customers_window, bg="white")
        export_frame.pack(fill=tk.X, padx=20, pady=10)

        def export_unique_customers():
            try:
                conn = connect()
                cur = conn.cursor()
                cur.execute("""
                    SELECT DISTINCT c.customerid, c.customer_name, c.phone, c.email, 
                           COUNT(s.saleid) as total_purchases,
                           SUM(s.totalprice) as total_spent,
                           MAX(s.saledate) as last_purchase
                    FROM customer c
                    JOIN sale s ON c.customerid = s.customerid
                    GROUP BY c.customerid, c.customer_name, c.phone, c.email
                    ORDER BY total_spent DESC
                """)

                data = cur.fetchall()
                headers = ["מספר לקוח", "שם לקוח", "טלפון", "אימייל",
                           "מספר רכישות", "סה\"כ הוצאות", "רכישה אחרונה"]

                file_path = self.excel_exporter.export_data_to_excel(
                    data, headers, "לקוחות ייחודיים עם פעילות"
                )

                if file_path:
                    messagebox.showinfo("הצלחה", f"הקובץ נשמר בהצלחה:\n{file_path}")
                    os.startfile(os.path.dirname(file_path))

                cur.close()
                conn.close()

            except Exception as e:
                messagebox.showerror("שגיאה", f"שגיאה ביצוא: {str(e)}")

        tk.Button(
            export_frame,
            text="📤 יצא לאקסל",
            command=export_unique_customers,
            bg="#059669",
            fg="white",
            font=("Segoe UI", 12, "bold"),
            width=15,
            height=2
        ).pack(side=tk.RIGHT)

        # טבלת נתונים
        tree_frame = tk.Frame(customers_window, bg="white")
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        columns = ("מספר לקוח", "שם לקוח", "טלפון", "מספר רכישות", "סה\"כ הוצאות", "רכישה אחרונה")
        tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=20)

        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=150, anchor="center")

        # צבעים ללקוחות VIP
        tree.tag_configure("vip", background="#f3e8ff", foreground="#7c3aed")
        tree.tag_configure("regular", background="#f8fafc", foreground="#374151")

        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)

        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # טעינת נתונים
        try:
            conn = connect()
            cur = conn.cursor()
            cur.execute("""
                SELECT DISTINCT c.customerid, c.customer_name, c.phone, 
                       COUNT(s.saleid) as total_purchases,
                       SUM(s.totalprice) as total_spent,
                       MAX(s.saledate) as last_purchase
                FROM customer c
                JOIN sale s ON c.customerid = s.customerid
                GROUP BY c.customerid, c.customer_name, c.phone
                ORDER BY total_spent DESC
            """)

            rows = cur.fetchall()
            for row in rows:
                total_spent = row[4]
                # קביעת צבע לפי סכום הוצאות
                tag = "vip" if total_spent >= 1000 else "regular"
                display_row = (row[0], row[1], row[2], row[3], f"₪{row[4]:,.0f}", row[5])
                tree.insert("", tk.END, values=display_row, tags=(tag,))

            cur.close()
            conn.close()

            # הצגת סיכום
            summary_frame = tk.Frame(customers_window, bg="#f8fafc", relief="solid", bd=1)
            summary_frame.pack(fill=tk.X, padx=20, pady=10)

            tk.Label(
                summary_frame,
                text=f"📊 סה\"כ נמצאו {len(rows)} לקוחות ייחודיים",
                font=("Segoe UI", 12, "bold"),
                bg="#f8fafc",
                fg="#8b5cf6"
            ).pack(pady=10)

        except Exception as e:
            tk.Label(tree_frame, text=f"שגיאה בטעינת נתונים: {e}", bg="white", fg="red").pack()

    def show_welcome_screen(self):
        """מסך ברירת מחדל"""
        self.clear_content()

        welcome_frame = tk.Frame(self.dynamic_content, bg="white")
        welcome_frame.pack(fill=tk.BOTH, expand=True, padx=50, pady=50)

        tk.Label(
            welcome_frame,
            text="🏪",
            font=("Segoe UI", 80),
            bg="white",
            fg="#3b82f6"
        ).pack(pady=(100, 30))

        tk.Label(
            welcome_frame,
            text="ברוכים הבאים למערכת ניהול החנות",
            font=("Segoe UI", 24, "bold"),
            bg="white",
            fg="#1e293b"
        ).pack(pady=20)

        tk.Label(
            welcome_frame,
            text="בחר אחת מהאפשרויות מהתפריט הימני כדי להתחיל",
            font=("Segoe UI", 16),
            bg="white",
            fg="#6b7280"
        ).pack(pady=10)

    def show_products_screen(self):
        """הצגת מסך ניהול מוצרים"""
        self.clear_content()

        # כותרת
        header = tk.Frame(self.dynamic_content, bg="#3b82f6", height=50)
        header.pack(fill=tk.X)
        header.pack_propagate(False)

        tk.Label(
            header,
            text="📦 ניהול מוצרים",
            font=("Segoe UI", 16, "bold"),
            bg="#3b82f6",
            fg="white"
        ).pack(expand=True)

        # תוכן עם גלילה
        canvas = tk.Canvas(self.dynamic_content, bg="white")
        scrollbar = ttk.Scrollbar(self.dynamic_content, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg="white")

        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # מסגרת עליונה לטופס וסטטיסטיקות
        top_frame = tk.Frame(scrollable_frame, bg="white")
        top_frame.pack(fill=tk.X, padx=20, pady=10)

        # טופס הוספת מוצר (צד שמאל)
        form_frame = tk.Frame(top_frame, bg="#f8fafc", relief="solid", bd=1)
        form_frame.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10))

        tk.Label(form_frame, text="🆕 הוספת מוצר חדש", font=("Segoe UI", 14, "bold"), bg="#f8fafc", fg="#1e293b").pack(pady=10)

        # שדות הטופס
        fields_frame = tk.Frame(form_frame, bg="#f8fafc")
        fields_frame.pack(padx=20, pady=10)

        def create_field(parent, label_text, width=15):
            row = tk.Frame(parent, bg="#f8fafc")
            row.pack(fill=tk.X, pady=3)
            tk.Label(row, text=label_text, bg="#f8fafc", font=("Segoe UI", 10, "bold"), width=12, anchor="e").pack(side=tk.RIGHT, padx=5)
            entry = tk.Entry(row, width=width, font=("Segoe UI", 10))
            entry.pack(side=tk.RIGHT, padx=5)
            return entry

        # הוספת שדה מספר מוצר לעדכון
        self.product_id_entry = create_field(fields_frame, "🔢 מספר מוצר:", 10)
        self.product_name_entry = create_field(fields_frame, "📝 שם מוצר:")
        self.product_price_entry = create_field(fields_frame, "💰 מחיר:", 10)
        self.product_amount_entry = create_field(fields_frame, "📊 כמות:", 10)
        self.product_category_entry = create_field(fields_frame, "🏷️ קטגוריה:")
        self.product_min_entry = create_field(fields_frame, "⚠️ מינימום:", 10)

        # כפתור הוספה
        def add_product():
            try:
                conn = connect()
                cur = conn.cursor()
                cur.execute("""
                    INSERT INTO product (product_name, price, amount, category, minamount, added_date, last_updated)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, (
                    self.product_name_entry.get(),
                    float(self.product_price_entry.get()),
                    int(self.product_amount_entry.get()),
                    self.product_category_entry.get(),
                    int(self.product_min_entry.get()),
                    dt.date.today(),
                    dt.date.today()
                ))
                conn.commit()
                cur.close()
                conn.close()
                messagebox.showinfo("הצלחה", "המוצר נוסף בהצלחה!")
                self.show_products_screen()
            except Exception as e:
                messagebox.showerror("שגיאה", str(e))

        # כפתורי פעולות מוצרים
        buttons_frame = tk.Frame(form_frame, bg="#f8fafc")
        buttons_frame.pack(pady=15)

        tk.Button(buttons_frame, text="➕ הוסף מוצר", command=add_product, bg="#3b82f6", fg="white", font=("Segoe UI", 10, "bold"), width=15, height=2).pack(side=tk.LEFT, padx=5)
        tk.Button(buttons_frame, text="✏️ עדכן מוצר", command=self.update_product, bg="#10b981", fg="white", font=("Segoe UI", 10, "bold"), width=15, height=2).pack(side=tk.LEFT, padx=5)
        tk.Button(buttons_frame, text="🗑️ מחק מוצר", command=self.delete_product, bg="#ef4444", fg="white", font=("Segoe UI", 10, "bold"), width=15, height=2).pack(side=tk.LEFT, padx=5)

        # סטטיסטיקות (צד ימין)
        stats_frame = tk.Frame(top_frame, bg="white")
        stats_frame.pack(side=tk.RIGHT, fill=tk.Y)

        try:
            conn = connect()
            cur = conn.cursor()

            cur.execute("SELECT COUNT(*) FROM product")
            total_products = cur.fetchone()[0]

            cur.execute("SELECT COUNT(*) FROM product WHERE amount < minamount")
            low_stock = cur.fetchone()[0]

            cur.execute("SELECT SUM(price * amount) FROM product")
            inventory_value = cur.fetchone()[0] or 0

            cur.close()
            conn.close()

            def create_vertical_stat_card(parent, title, value, icon, bg_color):
                card = tk.Frame(parent, bg=bg_color, width=200, height=90, relief="solid", bd=2)
                card.pack_propagate(False)
                card.pack(pady=10)

                top_row = tk.Frame(card, bg=bg_color)
                top_row.pack(fill=tk.X, pady=(10, 5))

                tk.Label(top_row, text=icon, font=("Segoe UI", 16), bg=bg_color).pack(side=tk.RIGHT, padx=5)
                tk.Label(top_row, text=title, font=("Segoe UI", 10, "bold"), bg=bg_color).pack(side=tk.RIGHT)

                tk.Label(card, text=value, font=("Segoe UI", 16, "bold"), bg=bg_color).pack()

            create_vertical_stat_card(stats_frame, "סה\"כ מוצרים", str(total_products), "📦", "#e3f2fd")
            create_vertical_stat_card(stats_frame, "מלאי נמוך", str(low_stock), "⚠️", "#fff8e1")
            create_vertical_stat_card(stats_frame, "שווי מלאי", f"₪{inventory_value:,.0f}", "💎", "#e8f5e9")

        except Exception as e:
            tk.Label(stats_frame, text=f"שגיאה: {e}", bg="white", fg="red").pack()

        # כפתורי יצוא לאקסל
        export_frame, buttons_container = self.create_export_buttons_frame(scrollable_frame)

        self.create_export_button(
            buttons_container,
            "יצא כל המוצרים",
            self.export_products_to_excel,
            "📦",
            "#059669"
        )

        self.create_export_button(
            buttons_container,
            "יצא מלאי נמוך",
            self.export_low_stock_to_excel,
            "⚠️",
            "#dc2626"
        )

        # טבלת מוצרים
        table_frame = tk.Frame(scrollable_frame, bg="white")
        table_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        tk.Label(table_frame, text="📋 רשימת מוצרים (10 האחרונים):", font=("Segoe UI", 14, "bold"), bg="white", fg="#1e293b").pack(anchor="e", pady=(0, 10))

        columns = ("מספר", "שם מוצר", "מחיר", "כמות", "קטגוריה", "מינימום")
        tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=10)

        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=120, anchor="center")

        tree_scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=tree_scrollbar.set)

        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        tree_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # טעינת 10 מוצרים אחרונים
        try:
            conn = connect()
            cur = conn.cursor()
            cur.execute("SELECT product_id, product_name, price, amount, category, minamount FROM product ORDER BY product_id DESC LIMIT 10")
            for row in cur.fetchall():
                tree.insert("", tk.END, values=row)
            cur.close()
            conn.close()
        except Exception as e:
            pass

    def show_sales_screen(self):
        """הצגת מסך ניהול מכירות"""
        self.clear_content()

        header = tk.Frame(self.dynamic_content, bg="#10b981", height=50)
        header.pack(fill=tk.X)
        header.pack_propagate(False)

        tk.Label(
            header,
            text="🛒 ניהול מכירות",
            font=("Segoe UI", 16, "bold"),
            bg="#10b981",
            fg="white"
        ).pack(expand=True)

        # תוכן עם גלילה
        canvas = tk.Canvas(self.dynamic_content, bg="white")
        scrollbar = ttk.Scrollbar(self.dynamic_content, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg="white")

        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # מסגרת עליונה לטופס וסטטיסטיקות
        top_frame = tk.Frame(scrollable_frame, bg="white")
        top_frame.pack(fill=tk.X, padx=20, pady=10)

        # טופס הוספת מכירה (צד שמאל)
        form_frame = tk.Frame(top_frame, bg="#f8fafc", relief="solid", bd=1)
        form_frame.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10))

        tk.Label(form_frame, text="🆕 הוספת מכירה חדשה", font=("Segoe UI", 14, "bold"), bg="#f8fafc", fg="#1e293b").pack(pady=10)

        # שדות הטופס
        fields_frame = tk.Frame(form_frame, bg="#f8fafc")
        fields_frame.pack(padx=20, pady=10)

        def create_sale_field(parent, label_text, icon, width=15):
            row = tk.Frame(parent, bg="#f8fafc")
            row.pack(fill=tk.X, pady=8)
            tk.Label(row, text=f"{icon} {label_text}", bg="#f8fafc", font=("Segoe UI", 10, "bold"), width=15, anchor="e").pack(side=tk.RIGHT, padx=5)
            entry = tk.Entry(row, width=width, font=("Segoe UI", 11), justify="center")
            entry.pack(side=tk.RIGHT, padx=5)
            return entry

        # הוספת שדה מספר מכירה לעדכון
        self.sale_id_entry = create_sale_field(fields_frame, "מספר מכירה (לעדכון)", "🔢")
        self.sale_total_entry = create_sale_field(fields_frame, "סכום כולל (₪)", "💰")
        self.sale_customer_entry = create_sale_field(fields_frame, "קוד לקוח", "👤")

        # כפתור הוספה
        def add_sale():
            try:
                conn = connect()
                cur = conn.cursor()
                cur.execute("""
                    INSERT INTO sale (saledate, totalprice, customerid)
                    VALUES (%s, %s, %s)
                """, (
                    dt.date.today(),
                    float(self.sale_total_entry.get()),
                    int(self.sale_customer_entry.get())
                ))
                conn.commit()
                cur.close()
                conn.close()
                messagebox.showinfo("הצלחה", "המכירה נוספה בהצלחה!")
                self.show_sales_screen()
            except Exception as e:
                messagebox.showerror("שגיאה", str(e))

        # כפתורי פעולות מכירות
        buttons_frame = tk.Frame(form_frame, bg="#f8fafc")
        buttons_frame.pack(pady=15)

        tk.Button(buttons_frame, text="➕ הוסף מכירה", command=add_sale, bg="#10b981", fg="white", font=("Segoe UI", 11, "bold"), width=18, height=2).pack(side=tk.LEFT, padx=5)
        tk.Button(buttons_frame, text="✏️ עדכן מכירה", command=self.update_sale, bg="#3b82f6", fg="white", font=("Segoe UI", 11, "bold"), width=18, height=2).pack(side=tk.LEFT, padx=5)
        tk.Button(buttons_frame, text="🗑️ מחק מכירה", command=self.delete_sale, bg="#ef4444", fg="white", font=("Segoe UI", 11, "bold"), width=18, height=2).pack(side=tk.LEFT, padx=5)

        # סטטיסטיקות (צד ימין)
        stats_frame = tk.Frame(top_frame, bg="white")
        stats_frame.pack(side=tk.RIGHT, fill=tk.Y)

        try:
            conn = connect()
            cur = conn.cursor()

            cur.execute("SELECT COUNT(*) FROM sale")
            total_sales = cur.fetchone()[0]

            cur.execute("SELECT SUM(totalprice) FROM sale")
            total_revenue = cur.fetchone()[0] or 0

            cur.execute("SELECT COUNT(*) FROM sale WHERE saledate = %s", (dt.date.today(),))
            today_sales = cur.fetchone()[0]

            cur.close()
            conn.close()

            def create_vertical_sale_card(parent, title, value, icon, bg_color):
                card = tk.Frame(parent, bg=bg_color, width=200, height=90, relief="solid", bd=2)
                card.pack_propagate(False)
                card.pack(pady=10)

                top_row = tk.Frame(card, bg=bg_color)
                top_row.pack(fill=tk.X, pady=(10, 5))

                tk.Label(top_row, text=icon, font=("Segoe UI", 16), bg=bg_color).pack(side=tk.RIGHT, padx=5)
                tk.Label(top_row, text=title, font=("Segoe UI", 10, "bold"), bg=bg_color).pack(side=tk.RIGHT)

                tk.Label(card, text=value, font=("Segoe UI", 16, "bold"), bg=bg_color).pack()

            create_vertical_sale_card(stats_frame, "סה\"כ מכירות", str(total_sales), "📊", "#dbeafe")
            create_vertical_sale_card(stats_frame, "הכנסות כוללות", f"₪{total_revenue:,.0f}", "💎", "#dcfce7")
            create_vertical_sale_card(stats_frame, "מכירות היום", str(today_sales), "🗓️", "#f3e8ff")

        except Exception as e:
            tk.Label(stats_frame, text=f"שגיאה: {e}", bg="white", fg="red").pack()

        # כפתורי יצוא לאקסל
        export_frame, buttons_container = self.create_export_buttons_frame(scrollable_frame)

        self.create_export_button(
            buttons_container,
            "יצא כל המכירות",
            self.export_sales_to_excel,
            "🛒",
            "#059669"
        )

        # טבלת מכירות
        table_frame = tk.Frame(scrollable_frame, bg="white")
        table_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        tk.Label(table_frame, text="📋 רשימת מכירות (10 האחרונות):", font=("Segoe UI", 14, "bold"), bg="white", fg="#1e293b").pack(anchor="e", pady=(0, 10))

        columns = ("מספר", "תאריך", "סכום", "לקוח")
        tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=10)

        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=150, anchor="center")

        tree_scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=tree_scrollbar.set)

        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        tree_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # טעינת 10 מכירות אחרונות
        try:
            conn = connect()
            cur = conn.cursor()
            cur.execute("SELECT saleid, saledate, totalprice, customerid FROM sale ORDER BY saleid DESC LIMIT 10")
            for row in cur.fetchall():
                tree.insert("", tk.END, values=(row[0], row[1], f"₪{row[2]:,.0f}", row[3]))
            cur.close()
            conn.close()
        except Exception as e:
            pass

    def show_stats_screen(self):
        """הצגת מסך סטטיסטיקות עם גרפים"""
        self.clear_content()

        header = tk.Frame(self.dynamic_content, bg="#8b5cf6", height=50)
        header.pack(fill=tk.X)
        header.pack_propagate(False)

        tk.Label(
            header,
            text="📊 סטטיסטיקות החנות",
            font=("Segoe UI", 16, "bold"),
            bg="#8b5cf6",
            fg="white"
        ).pack(expand=True)

        # תוכן עם גלילה
        canvas = tk.Canvas(self.dynamic_content, bg="white")
        scrollbar = ttk.Scrollbar(self.dynamic_content, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg="white")

        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # כרטיסי סטטיסטיקות עליונים
        top_stats_frame = tk.Frame(scrollable_frame, bg="white")
        top_stats_frame.pack(fill=tk.X, padx=20, pady=20)

        try:
            conn = connect()
            cur = conn.cursor()

            # נתונים בסיסיים
            cur.execute("SELECT COUNT(*) FROM sale")
            total_sales = cur.fetchone()[0]

            cur.execute("SELECT SUM(totalprice) FROM sale")
            total_revenue = cur.fetchone()[0] or 0

            cur.execute("SELECT COUNT(*) FROM product WHERE amount < minamount")
            low_stock = cur.fetchone()[0]

            cur.execute("SELECT COUNT(DISTINCT customerid) FROM sale")
            unique_customers = cur.fetchone()[0]

            # יצירת כרטיסים אינטראקטיביים
            def create_clickable_stat_card(parent, title, value, icon, bg_color, command=None):
                card = tk.Frame(parent, bg=bg_color, width=220, height=100, relief="solid", bd=2, cursor="hand2" if command else "arrow")
                card.pack_propagate(False)
                card.pack(side="right", padx=15, pady=10)

                if command:
                    card.bind("<Button-1>", lambda e: command())

                top_row = tk.Frame(card, bg=bg_color)
                top_row.pack(fill=tk.X, pady=(15, 5))

                tk.Label(top_row, text=icon, font=("Segoe UI", 20), bg=bg_color).pack(side=tk.RIGHT, padx=10)
                tk.Label(top_row, text=title, font=("Segoe UI", 11, "bold"), bg=bg_color).pack(side=tk.RIGHT)

                tk.Label(card, text=value, font=("Segoe UI", 18, "bold"), bg=bg_color).pack()

                if command:
                    tk.Label(card, text="👆 לחץ לפרטים", font=("Segoe UI", 8), bg=bg_color, fg="#6b7280").pack()

                return card

            # פונקציות לכפתורים
            def show_sales_details():
                self.show_detail_window("מכירות", "SELECT saleid, saledate, totalprice, customerid FROM sale ORDER BY saleid DESC LIMIT 50",
                                        ["מספר מכירה", "תאריך", "סכום", "לקוח"])

            def show_low_stock_details():
                self.show_low_stock_window()

            def show_unique_customers_details():
                self.show_unique_customers_window()

            create_clickable_stat_card(top_stats_frame, "סה\"כ מכירות", str(total_sales), "🛒", "#dbeafe", show_sales_details)
            create_clickable_stat_card(top_stats_frame, "הכנסות כוללות", f"₪{total_revenue:,.0f}", "💰", "#dcfce7")
            create_clickable_stat_card(top_stats_frame, "מלאי נמוך", str(low_stock), "⚠️", "#fef3c7", show_low_stock_details)
            create_clickable_stat_card(top_stats_frame, "לקוחות ייחודיים", str(unique_customers), "👥", "#f3e8ff", show_unique_customers_details)

            # גרף מכירות לפי יום
            chart_frame = tk.Frame(scrollable_frame, bg="#f8fafc", relief="solid", bd=2)
            chart_frame.pack(fill=tk.X, padx=20, pady=20)

            tk.Label(chart_frame, text="📈 מכירות השבוע", font=("Segoe UI", 16, "bold"), bg="#f8fafc", fg="#1e293b").pack(pady=15)

            # יצירת גרף פשוט עם טקסט
            days = ["ראשון", "שני", "שלישי", "רביעי", "חמישי", "שישי", "שבת"]
            sales_data = [12, 18, 15, 22, 25, 30, 8]

            graph_frame = tk.Frame(chart_frame, bg="white", relief="solid", bd=1)
            graph_frame.pack(fill=tk.X, padx=20, pady=20)

            for i, (day, sales) in enumerate(zip(days, sales_data)):
                day_frame = tk.Frame(graph_frame, bg="white")
                day_frame.pack(side=tk.RIGHT, padx=10, pady=10)

                bar_height = max(10, sales * 3)
                bar = tk.Frame(day_frame, bg="#3b82f6", width=40, height=bar_height)
                bar.pack()
                bar.pack_propagate(False)

                tk.Label(day_frame, text=str(sales), font=("Segoe UI", 10, "bold"), bg="white").pack(pady=2)
                tk.Label(day_frame, text=day, font=("Segoe UI", 9), bg="white").pack()

            cur.close()
            conn.close()

        except Exception as e:
            tk.Label(scrollable_frame, text=f"שגיאה בטעינת נתונים: {e}", bg="white", fg="red").pack()

    def show_detail_window(self, title, query, columns):
        """הצגת חלון פרטים"""
        detail_window = tk.Toplevel(self.root)
        detail_window.title(f"פרטי {title}")
        detail_window.geometry("800x500")
        detail_window.configure(bg="white")

        tk.Label(detail_window, text=f"📋 פרטי {title}", font=("Segoe UI", 16, "bold"), bg="white", fg="#1e293b").pack(pady=20)

        tree = ttk.Treeview(detail_window, columns=columns, show="headings", height=20)

        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=150, anchor="center")

        scrollbar = ttk.Scrollbar(detail_window, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)

        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=20, pady=20)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y, pady=20)

        try:
            conn = connect()
            cur = conn.cursor()
            cur.execute(query)
            for row in cur.fetchall():
                tree.insert("", tk.END, values=row)
            cur.close()
            conn.close()
        except Exception as e:
            tk.Label(detail_window, text=f"שגיאה: {e}", bg="white", fg="red").pack()

    def show_discounts_screen(self):
        """הצגת מסך ניהול הנחות"""
        self.clear_content()

        header = tk.Frame(self.dynamic_content, bg="#f59e0b", height=50)
        header.pack(fill=tk.X)
        header.pack_propagate(False)

        tk.Label(
            header,
            text="💸 ניהול הנחות",
            font=("Segoe UI", 16, "bold"),
            bg="#f59e0b",
            fg="white"
        ).pack(expand=True)

        # תוכן עם גלילה
        canvas = tk.Canvas(self.dynamic_content, bg="white")
        scrollbar = ttk.Scrollbar(self.dynamic_content, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg="white")

        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # טופס הוספת הנחה
        form_frame = tk.Frame(scrollable_frame, bg="#f8fafc", relief="solid", bd=2)
        form_frame.pack(fill=tk.X, padx=20, pady=10)

        tk.Label(form_frame, text="🆕 הוספת הנחה חדשה", font=("Segoe UI", 16, "bold"), bg="#f8fafc", fg="#1e293b").pack(pady=15)

        # שדות הטופס
        fields_container = tk.Frame(form_frame, bg="#f8fafc")
        fields_container.pack(pady=20)

        row1 = tk.Frame(fields_container, bg="#f8fafc")
        row1.pack(pady=10)

        def create_discount_field(parent, label_text, icon, width=12):
            field_frame = tk.Frame(parent, bg="#f8fafc")
            field_frame.pack(side=tk.RIGHT, padx=15)
            tk.Label(field_frame, text=f"{icon} {label_text}", bg="#f8fafc", font=("Segoe UI", 10, "bold")).pack()
            entry = tk.Entry(field_frame, width=width, font=("Segoe UI", 10), justify="center")
            entry.pack(pady=5)
            return entry

        self.discount_id_entry = create_discount_field(row1, "מספר הנחה (לעדכון)", "🔢", 10)
        self.discount_rate_entry = create_discount_field(row1, "שיעור הנחה (%)", "📊", 10)
        self.discount_product_entry = create_discount_field(row1, "מספר מוצר", "📦", 10)

        row2 = tk.Frame(fields_container, bg="#f8fafc")
        row2.pack(pady=10)

        self.discount_store_entry = create_discount_field(row2, "מספר סניף", "🏪", 10)
        self.discount_start_entry = create_discount_field(row2, "תאריך התחלה", "📅", 15)
        self.discount_end_entry = create_discount_field(row2, "תאריך סיום", "📅", 15)

        tk.Label(fields_container, text="💡 פורמט תאריך: YYYY-MM-DD (לדוגמה: 2024-12-25)",
                 bg="#f8fafc", font=("Segoe UI", 9), fg="#6b7280").pack(pady=5)

        # כפתור הוספה
        def add_discount():
            try:
                conn = connect()
                cur = conn.cursor()
                cur.execute("""
                    INSERT INTO discount (discountrate, startdate, enddate, storeid, productid)
                    VALUES (%s, %s, %s, %s, %s)
                """, (
                    float(self.discount_rate_entry.get()),
                    self.discount_start_entry.get(),
                    self.discount_end_entry.get(),
                    int(self.discount_store_entry.get()),
                    int(self.discount_product_entry.get())
                ))
                conn.commit()
                cur.close()
                conn.close()
                messagebox.showinfo("הצלחה", "ההנחה נוספה בהצלחה!")
                # ניקוי השדות
                for entry in [self.discount_rate_entry, self.discount_product_entry, self.discount_store_entry,
                              self.discount_start_entry, self.discount_end_entry]:
                    entry.delete(0, tk.END)
                self.show_discounts_screen()
            except Exception as e:
                messagebox.showerror("שגיאה", str(e))

        # כפתורי פעולות הנחות
        buttons_frame = tk.Frame(form_frame, bg="#f8fafc")
        buttons_frame.pack(pady=20)

        tk.Button(buttons_frame, text="➕ הוסף הנחה", command=add_discount, bg="#f59e0b", fg="white", font=("Segoe UI", 11, "bold"), width=18, height=2).pack(side=tk.LEFT, padx=5)
        tk.Button(buttons_frame, text="✏️ עדכן הנחה", command=self.update_discount, bg="#3b82f6", fg="white", font=("Segoe UI", 11, "bold"), width=18, height=2).pack(side=tk.LEFT, padx=5)
        tk.Button(buttons_frame, text="🗑️ מחק הנחה", command=self.delete_discount, bg="#ef4444", fg="white", font=("Segoe UI", 11, "bold"), width=18, height=2).pack(side=tk.LEFT, padx=5)

        # כפתורי יצוא לאקסל
        export_frame, buttons_container = self.create_export_buttons_frame(scrollable_frame)

        self.create_export_button(
            buttons_container,
            "יצא רשימת הנחות",
            self.export_discounts_to_excel,
            "💸",
            "#059669"
        )

        # טבלת הנחות
        table_frame = tk.Frame(scrollable_frame, bg="white")
        table_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        tk.Label(table_frame, text="📋 רשימת הנחות (10 האחרונות):", font=("Segoe UI", 14, "bold"), bg="white", fg="#1e293b").pack(anchor="e", pady=(0, 10))

        columns = ("מספר", "מוצר", "סניף", "הנחה %", "תחילה", "סיום", "סטטוס")
        tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=10)

        for i, col in enumerate(columns):
            tree.heading(col, text=col)
            width = 80 if col == "מספר" else 120 if col in ["הנחה %", "תחילה", "סיום", "סטטוס"] else 150
            tree.column(col, width=width, anchor="center")

        # צבעים לסטטוס
        tree.tag_configure("active", background="#dcfce7", foreground="#15803d")
        tree.tag_configure("inactive", background="#fef2f2", foreground="#dc2626")

        tree_scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=tree_scrollbar.set)

        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        tree_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # טעינת נתונים עם בדיקת סטטוס
        try:
            conn = connect()
            cur = conn.cursor()
            cur.execute("""
                SELECT d.discountid, p.product_name, s.storelocation, d.discountrate, d.startdate, d.enddate
                FROM discount d
                JOIN product p ON d.productid = p.product_id
                JOIN store s ON d.storeid = s.storeid
                ORDER BY d.discountid DESC LIMIT 10
            """)

            for row in cur.fetchall():
                from datetime import datetime as dt_check
                start_date = dt_check.strptime(str(row[4]), "%Y-%m-%d")
                end_date = dt_check.strptime(str(row[5]), "%Y-%m-%d")
                today = dt_check.now()

                if start_date <= today <= end_date:
                    status = "פעילה ✅"
                    tag = "active"
                else:
                    status = "לא פעילה ❌"
                    tag = "inactive"

                tree.insert("", tk.END, values=(
                    row[0], row[1], row[2], f"{row[3]:.1f}%", row[4], row[5], status
                ), tags=(tag,))

            cur.close()
            conn.close()
        except Exception as e:
            tk.Label(table_frame, text=f"שגיאה: {e}", bg="white", fg="red").pack()

    def show_queries_screen(self):
        """הצגת מסך שאילתות"""
        self.clear_content()

        header = tk.Frame(self.dynamic_content, bg="#6366f1", height=50)
        header.pack(fill=tk.X)
        header.pack_propagate(False)

        tk.Label(
            header,
            text="⚙️ שאילתות ופרוצדורות",
            font=("Segoe UI", 16, "bold"),
            bg="#6366f1",
            fg="white"
        ).pack(expand=True)

        # תוכן עם גלילה
        canvas = tk.Canvas(self.dynamic_content, bg="white")
        scrollbar = ttk.Scrollbar(self.dynamic_content, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg="white")

        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # בחירת שאילתה
        selection_frame = tk.Frame(scrollable_frame, bg="#f8fafc", relief="solid", bd=2)
        selection_frame.pack(fill=tk.X, padx=20, pady=10)

        tk.Label(selection_frame, text="🔍 בחר שאילתה להפעלה", font=("Segoe UI", 16, "bold"), bg="#f8fafc", fg="#1e293b").pack(pady=15)

        # מילון שאילתות
        self.queries_data = {
            "כל המוצרים": {
                "query": "SELECT product_id, product_name, price, amount, category FROM product ORDER BY product_name",
                "description": "הצגת כל המוצרים במערכת",
                "icon": "📦",
                "params": []
            },
            "מוצרים עם מלאי נמוך": {
                "query": "SELECT product_id, product_name, amount, minamount FROM product WHERE amount < minamount",
                "description": "מוצרים שהכמות שלהם נמוכה מהמינימום הנדרש",
                "icon": "⚠️",
                "params": []
            },
            "כל המכירות": {
                "query": "SELECT saleid, saledate, totalprice, customerid FROM sale ORDER BY saledate DESC",
                "description": "רשימת כל המכירות במערכת",
                "icon": "🛒",
                "params": []
            },
            "מכירות לפי לקוח": {
                "query": "SELECT s.saleid, s.saledate, s.totalprice FROM sale s WHERE s.customerid = %s ORDER BY s.saledate DESC",
                "description": "מכירות של לקוח ספציפי",
                "icon": "👤",
                "params": [("קוד לקוח", "int")]
            },
            "הנחות פעילות": {
                "query": "SELECT d.discountid, p.product_name, d.discountrate FROM discount d JOIN product p ON d.productid = p.product_id WHERE d.startdate <= CURRENT_DATE AND d.enddate >= CURRENT_DATE",
                "description": "הנחות שפעילות כרגע",
                "icon": "💸",
                "params": []
            }
        }

        # בחירת שאילתה
        query_selection_frame = tk.Frame(selection_frame, bg="#f8fafc")
        query_selection_frame.pack(fill=tk.X, padx=20, pady=10)

        tk.Label(query_selection_frame, text="בחר שאילתה:", font=("Segoe UI", 12, "bold"), bg="#f8fafc").pack(anchor="e", pady=5)

        self.selected_query = tk.StringVar()
        self.query_combo = ttk.Combobox(query_selection_frame, textvariable=self.selected_query,
                                        values=list(self.queries_data.keys()), state="readonly",
                                        font=("Segoe UI", 11), width=40)
        self.query_combo.pack(anchor="e", pady=5)
        self.query_combo.bind("<<ComboboxSelected>>", self.on_query_selected)

        # תיאור השאילתה
        self.description_label = tk.Label(selection_frame, text="", font=("Segoe UI", 10),
                                          bg="#f8fafc", fg="#6b7280", wraplength=600, justify="right")
        self.description_label.pack(anchor="e", padx=20, pady=10)

        # פרמטרים
        self.params_frame = tk.Frame(scrollable_frame, bg="#e0f2fe", relief="solid", bd=2)
        self.params_frame.pack(fill=tk.X, padx=20, pady=10)

        tk.Label(self.params_frame, text="📝 פרמטרים", font=("Segoe UI", 14, "bold"), bg="#e0f2fe", fg="#1e293b").pack(pady=10)

        self.param_entries = {}
        self.param_widgets_frame = tk.Frame(self.params_frame, bg="#e0f2fe")
        self.param_widgets_frame.pack(padx=20, pady=10)

        # כפתור הפעלה
        execute_frame = tk.Frame(scrollable_frame, bg="white")
        execute_frame.pack(fill=tk.X, padx=20, pady=20)

        self.execute_btn = tk.Button(
            execute_frame,
            text="▶️ הפעל שאילתה",
            command=self.execute_selected_query,
            font=("Segoe UI", 14, "bold"),
            bg="#10b981",
            fg="white",
            width=25,
            height=2,
            relief="flat",
            cursor="hand2",
            state="disabled"
        )
        self.execute_btn.pack()

        # כפתור יצוא תוצאות (מוסתר בהתחלה)
        self.export_results_btn = tk.Button(
            execute_frame,
            text="📤 יצא תוצאות לאקסל",
            command=self.export_query_results,
            font=("Segoe UI", 12, "bold"),
            bg="#059669",
            fg="white",
            width=25,
            height=2,
            relief="flat",
            cursor="hand2"
        )

        # תוצאות
        results_frame = tk.Frame(scrollable_frame, bg="white")
        results_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        tk.Label(results_frame, text="📊 תוצאות:", font=("Segoe UI", 14, "bold"), bg="white", fg="#1e293b").pack(anchor="e", pady=(0, 10))

        self.results_tree = ttk.Treeview(results_frame, show="headings", height=12)
        results_scrollbar = ttk.Scrollbar(results_frame, orient=tk.VERTICAL, command=self.results_tree.yview)
        self.results_tree.configure(yscrollcommand=results_scrollbar.set)

        self.results_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        results_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # סטטוס
        self.status_label = tk.Label(scrollable_frame, text="בחר שאילתה כדי להתחיל",
                                     font=("Segoe UI", 11), bg="white", fg="#6b7280")
        self.status_label.pack(pady=10)

    def on_query_selected(self, event=None):
        """טיפול בבחירת שאילתה"""
        query_name = self.selected_query.get()
        if not query_name:
            return

        query_data = self.queries_data[query_name]

        # עדכון תיאור
        self.description_label.config(text=f"{query_data['icon']} {query_data['description']}")

        # ניקוי פרמטרים קיימים
        for widget in self.param_widgets_frame.winfo_children():
            widget.destroy()
        self.param_entries.clear()

        # יצירת פרמטרים חדשים
        params = query_data["params"]
        if params:
            for param_name, param_type in params:
                param_row = tk.Frame(self.param_widgets_frame, bg="#e0f2fe")
                param_row.pack(fill=tk.X, pady=5)

                tk.Label(param_row, text=f"{param_name}:", font=("Segoe UI", 11, "bold"),
                         bg="#e0f2fe", width=20, anchor="e").pack(side=tk.RIGHT, padx=5)

                entry = tk.Entry(param_row, width=20, font=("Segoe UI", 11), justify="center")
                entry.pack(side=tk.RIGHT, padx=5)

                self.param_entries[param_name] = (entry, param_type)
        else:
            tk.Label(self.param_widgets_frame, text="📌 אין פרמטרים נדרשים לשאילתה זו",
                     font=("Segoe UI", 10), bg="#e0f2fe", fg="#6b7280").pack(pady=10)

        # הפעלת כפתור הביצוע
        self.execute_btn.config(state="normal", bg="#10b981")
        self.status_label.config(text="מוכן להפעלה", fg="#10b981")

        # הסתרת כפתור יצוא תוצאות
        self.export_results_btn.pack_forget()

    def execute_selected_query(self):
        """הפעלת השאילתה הנבחרת"""
        query_name = self.selected_query.get()
        if not query_name:
            messagebox.showerror("שגיאה", "בחר שאילתה להפעלה")
            return

        query_data = self.queries_data[query_name]
        query = query_data["query"]
        params = query_data["params"]

        # איסוף פרמטרים
        param_values = []
        try:
            for param_name, param_type in params:
                entry, expected_type = self.param_entries[param_name]
                value = entry.get().strip()

                if not value:
                    raise ValueError(f"חסר ערך לפרמטר: {param_name}")

                if expected_type == "int":
                    value = int(value)
                elif expected_type == "float":
                    value = float(value)

                param_values.append(value)
        except ValueError as e:
            messagebox.showerror("שגיאה בפרמטרים", str(e))
            return

        # ביצוע השאילתה
        try:
            self.status_label.config(text="מבצע שאילתה...", fg="#f59e0b")

            conn = connect()
            cur = conn.cursor()

            if param_values:
                cur.execute(query, tuple(param_values))
            else:
                cur.execute(query)

            results = cur.fetchall()
            column_names = [desc[0] for desc in cur.description] if cur.description else []

            # שמירת התוצאות לייצוא
            self.last_query_results = results
            self.last_query_headers = column_names
            self.last_query_title = query_name

            # ניקוי תוצאות קודמות
            for item in self.results_tree.get_children():
                self.results_tree.delete(item)

            # הגדרת עמודות
            self.results_tree["columns"] = column_names
            for col in column_names:
                self.results_tree.heading(col, text=col)
                self.results_tree.column(col, width=120, anchor="center")

            # הוספת תוצאות
            for row in results:
                self.results_tree.insert("", tk.END, values=row)

            cur.close()
            conn.close()

            # עדכון סטטוס והצגת כפתור יצוא
            if len(results) > 0:
                self.status_label.config(text=f"✅ נמצאו {len(results)} תוצאות", fg="#10b981")
                self.export_results_btn.pack(pady=10)
            else:
                self.status_label.config(text="ℹ️ לא נמצאו תוצאות", fg="#6b7280")
                self.export_results_btn.pack_forget()

        except Exception as e:
            messagebox.showerror("שגיאה", str(e))
            self.status_label.config(text=f"❌ שגיאה: {str(e)}", fg="#ef4444")
            self.export_results_btn.pack_forget()

    def export_query_results(self):
        """יצוא תוצאות השאילתה הנוכחית לאקסל"""
        if not hasattr(self, 'last_query_results') or not self.last_query_results:
            messagebox.showwarning("אזהרה", "אין תוצאות לייצוא")
            return

        try:
            file_path = self.excel_exporter.export_data_to_excel(
                self.last_query_results,
                self.last_query_headers,
                f"תוצאות שאילתה - {self.last_query_title}"
            )

            if file_path:
                messagebox.showinfo("הצלחה", f"התוצאות יוצאו בהצלחה:\n{file_path}")
                os.startfile(os.path.dirname(file_path))

        except Exception as e:
            messagebox.showerror("שגיאה", f"שגיאה ביצוא התוצאות: {str(e)}")

# יצירת האפליקציה הראשית
def create_main_app():
    root = tk.Tk()
    app = MainApplication(root)

    # הודעת סטטוס בתחתית
    status_frame = tk.Frame(root, bg=BG_COLOR)
    status_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=10)

    status_label = tk.Label(
        status_frame,
        text="● מערכת פעילה | תכונת יצוא לאקסל זמינה",
        font=("Segoe UI", 12),
        fg="#10b981",
        bg=BG_COLOR
    )
    status_label.pack()

    root.mainloop()

# הפעלה
if __name__ == "__main__":
    create_main_app()
